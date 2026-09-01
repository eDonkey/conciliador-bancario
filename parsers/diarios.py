# -*- coding: utf-8 -*-
"""Parsers de los extractos diarios multibanco del grupo Nave y de los
reportes del sistema contable FBS (mayor E y O por cuenta).

`identificar(nombre, data)` detecta solo qué es cada archivo:
  - extracto de Santander/Río ("xls" que en realidad es texto tabulado),
    BBVA/Francés (.xls binario), Galicia (.xlsx), Macro (.xls binario)
    o Ciudad (.csv con ';'),
  - o un mayor FBS (.xls binario, hoja "(E) (codigo)" / "(O) (codigo)").

Devuelve la metadata + los movimientos/asientos ya parseados. La cuenta se
reporta como los dígitos detectados dentro del archivo (Galicia no trae
ninguno: queda para asignación manual).
"""
import html
import io
import re
from datetime import date, datetime, timedelta

import xlrd

from parsers.santander_pdf import MovimientoBanco
from parsers.mayor_xlsx import AsientoMayor

BASE_EXCEL = date(1899, 12, 30)
FBS_HOJA_RE = re.compile(r'\((E|O)\)\s*\((\d+)\)')


def _digits(s: str) -> str:
    return re.sub(r'\D', '', s or '')


def _importe_ar(s: str) -> float:
    """'1.850.205,68' / '(467.935,00)' / '-26107,00' -> float con signo."""
    s = (s or '').strip()
    neg = s.startswith('(') or s.startswith('-')
    v = s.strip('()').lstrip('-').replace('.', '').replace(',', '.')
    return -float(v) if neg and v else (float(v) if v else 0.0)


def _fecha(s):
    """Fecha desde 'dd/mm/yyyy', 'dd-mm-yyyy', datetime o serial de Excel."""
    if isinstance(s, datetime):
        return s.date()
    if isinstance(s, date):
        return s
    if isinstance(s, float) and 20000 < s < 80000:   # serial Excel plausible
        return BASE_EXCEL + timedelta(days=int(s))
    m = re.match(r'(\d{2})[/-](\d{2})[/-](\d{4})', str(s or ''))
    if m:
        return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
    m = re.match(r'(\d{4})-(\d{2})-(\d{2})', str(s or ''))
    if m:
        return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    return None


def _mov(seq, archivo, fecha, descripcion, importe, saldo=0.0,
         comprobante=None, detalle=""):
    return MovimientoBanco(
        id=f"B#{seq}", fecha=fecha, comprobante=comprobante,
        descripcion=descripcion.strip(), detalle=(detalle or '').strip(),
        debito=round(-importe, 2) if importe < 0 else 0.0,
        credito=round(importe, 2) if importe > 0 else 0.0,
        saldo=saldo, archivo=archivo, pagina=0)


# --- Santander / Río: texto tabulado disfrazado de .xls ---------------------

def _es_rio(data: bytes) -> bool:
    inicio = data[:400].decode('latin-1', errors='replace')
    return ('\t' in inicio or 'Movimientos' in inicio) and \
        data[:2] in (b'\r\n', b'\n\r') and b'\tConcepto\t' in data[:2000].replace(b'\r', b'')


def _parse_rio(nombre, data):
    texto = html.unescape(data.decode('latin-1', errors='replace'))
    m = re.search(r'Cuenta Corriente en (Pesos|D[oó]lares)\s+Nro\.\s*([\d/-]+)', texto)
    moneda = 'USD' if m and 'lares' in m.group(1) else 'ARS'
    cuenta = _digits(m.group(2)) if m else None
    movs, seq = [], 0
    for linea in texto.splitlines():
        c = linea.split('\t')
        if len(c) < 8 or not re.match(r'\d{2}/\d{2}/\d{4}', c[0]):
            continue
        seq += 1
        desc, _, det = c[5].partition('  - ')
        movs.append(_mov(seq, nombre, _fecha(c[0]), desc, _importe_ar(c[6]),
                         _importe_ar(c[7]), comprobante=c[4].strip() or None,
                         detalle=det))
    return {"banco": "santander", "cuenta": cuenta, "moneda": moneda,
            "movimientos": movs}


# --- Ciudad: CSV latin-1 con ';' --------------------------------------------

def _parse_ciudad(nombre, data):
    texto = data.decode('latin-1', errors='replace')
    movs, seq = [], 0
    cuenta = moneda = None
    for linea in texto.splitlines()[1:]:
        c = [x.strip() for x in linea.split(';')]
        if len(c) < 7 or not re.match(r'\d{2}/\d{2}/\d{4}', c[2]):
            continue
        if cuenta is None:
            cuenta = _digits(c[0]).lstrip('0')
            moneda = 'USD' if 'U$S' in c[0].upper() else 'ARS'
        seq += 1
        movs.append(_mov(seq, nombre, _fecha(c[2]), c[5], _importe_ar(c[3]),
                         _importe_ar(c[6]), comprobante=c[4] or None))
    return {"banco": "ciudad", "cuenta": cuenta, "moneda": moneda,
            "movimientos": movs}


# --- Galicia: .xlsx, hoja 'Movimientos' (sin número de cuenta adentro) ------

def _parse_galicia(nombre, data):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb['Movimientos'] if 'Movimientos' in wb.sheetnames else wb.worksheets[0]
    filas = list(ws.iter_rows(values_only=True))
    hdr = {str(v).strip(): i for i, v in enumerate(filas[0]) if v}
    movs = []
    for seq, f in enumerate(filas[1:], start=1):
        fecha = _fecha(f[hdr['Fecha']])
        if fecha is None:
            continue
        deb = float(f[hdr['Débitos']] or 0)
        cre = float(f[hdr['Créditos']] or 0)
        detalle = ' '.join(str(f[hdr[k]]).strip() for k in
                           ('Concepto', 'Leyendas Adicionales 1', 'Leyendas Adicionales 2')
                           if k in hdr and f[hdr[k]])
        movs.append(_mov(seq, nombre, fecha, str(f[hdr['Descripción']] or ''),
                         cre - deb, float(f[hdr['Saldo']] or 0),
                         comprobante=str(f[hdr['Número de Comprobante']] or '') or None,
                         detalle=detalle))
    return {"banco": "galicia", "cuenta": None, "moneda": None,
            "movimientos": movs}


# --- BBVA / Francés: .xls binario -------------------------------------------

def _parse_bbva(nombre, wb):
    cuenta = moneda = None
    movs = []
    seq = 0
    for hoja in wb.sheets():
        filas = [[hoja.cell_value(r, c) for c in range(hoja.ncols)]
                 for r in range(hoja.nrows)]
        hdr_i = None
        for i, f in enumerate(filas):
            if str(f[0]).startswith('Cuenta:') and len(f) > 1:
                m = re.search(r'([\d/-]+)\s*\(CC\s*(U\$S|\$)\)', str(f[1]))
                if m:
                    cuenta = _digits(m.group(1))
                    moneda = 'USD' if m.group(2) == 'U$S' else 'ARS'
            if f[0] == 'Fecha' and 'Concepto' in f:
                hdr_i = i
                cols = {str(v).strip(): j for j, v in enumerate(f) if str(v).strip()}
                continue
            if hdr_i is None or not _fecha(f[0]):
                continue
            fecha = _fecha(f[0])
            cre = abs(float(f[cols['Crédito']] or 0)) if f[cols['Crédito']] != '' else 0.0
            deb = abs(float(f[cols['Débito']] or 0)) if f[cols['Débito']] != '' else 0.0
            col_comp = next((cols[k] for k in ('Nro de cheque', 'Número Documento')
                             if k in cols), None)
            comp = str(f[col_comp] or '') if col_comp is not None else ''
            detalle = str(f[cols['Detalle']] or '') if 'Detalle' in cols else ''
            # No se deduplica: filas idénticas (p. ej. dos impuestos iguales el
            # mismo día) son operaciones reales y se conservan todas.
            seq += 1
            movs.append(_mov(seq, nombre, fecha, str(f[cols['Concepto']]),
                             cre - deb, comprobante=comp.strip() or None,
                             detalle=detalle))
    return {"banco": "frances", "cuenta": cuenta, "moneda": moneda,
            "movimientos": movs}


# --- Macro: .xls binario ----------------------------------------------------

def _parse_macro(nombre, wb):
    hoja = wb.sheets()[0]
    filas = [[hoja.cell_value(r, c) for c in range(hoja.ncols)]
             for r in range(hoja.nrows)]
    cuenta = moneda = None
    cols = None
    movs, seq = [], 0
    for f in filas:
        celdas = [str(v).strip() for v in f]
        if celdas[0] == 'Número':
            cuenta = _digits(next((v for v in celdas[1:] if _digits(v)), ''))
        elif celdas[0] == 'Moneda':
            moneda = 'USD' if any('DOLAR' in v.upper() for v in celdas[1:]) else 'ARS'
        elif celdas[0] == 'Fecha' and 'Causal' in celdas:
            cols = {v: i for i, v in enumerate(celdas) if v}
        elif cols and isinstance(f[0], float):
            seq += 1
            movs.append(_mov(seq, nombre, _fecha(f[0]),
                             str(f[cols['Concepto']] or ''),
                             float(f[cols['Importe']] or 0),
                             float(f[cols['Saldo']] or 0),
                             comprobante=str(f[cols['Nro. de Referencia']] or '') or None,
                             detalle=str(f[cols['Causal']] or '')))
    return {"banco": "macro", "cuenta": cuenta, "moneda": moneda,
            "movimientos": movs}


# --- Santander exportado a .xlsx (mismas columnas que el texto tabulado) ----

CUENTA_TXT_RE = re.compile(r'\b(\d{3})[-. ](\d{1,6})[-/.](\d)\b')


def _parse_rio_xlsx(nombre, wb):
    ws = wb.worksheets[0]
    filas = list(ws.iter_rows(values_only=True))
    moneda = None
    hdr = None
    movs, seq = [], 0
    for f in filas:
        c0 = str(f[0] or '')
        if 'Cuenta corriente en' in c0 or 'Cuenta Corriente en' in c0:
            moneda = 'USD' if 'lares' in c0 or 'U$S' in c0.upper() else 'ARS'
        elif c0 == 'Fecha' and 'Cod. Operativo' in [str(x).strip() for x in f]:
            hdr = {str(x).strip(): i for i, x in enumerate(f) if x}
        elif hdr and _fecha(f[0]) and not c0.startswith('Saldo'):
            seq += 1
            desc, _, det = str(f[hdr['Concepto']] or '').partition('  - ')
            saldo = f[hdr.get('Saldo', -1)] if 'Saldo' in hdr else 0
            movs.append(_mov(seq, nombre, _fecha(f[hdr['Fecha']]), desc,
                             float(str(f[hdr['Importe']] or 0).replace(',', '')),
                             float(str(saldo or 0).replace(',', '') or 0),
                             comprobante=str(f[hdr['Referencia']] or '') or None))
    # el número de cuenta no viene adentro: está en el nombre de hoja/archivo
    m = CUENTA_TXT_RE.search(f"{ws.title} {nombre}")
    cuenta = '-'.join(m.groups()) if m else None
    if moneda is None and re.search(r'CC\s*\$', f"{ws.title} {nombre}"):
        moneda = 'ARS'
    return {"banco": "santander", "cuenta": cuenta, "moneda": moneda,
            "movimientos": movs}


# --- FBS: Consulta del Mayor (hoja E u O) -----------------------------------

def _parse_fbs(nombre, wb):
    hoja = wb.sheets()[0]
    m = FBS_HOJA_RE.search(hoja.name)
    letra = m.group(1) if m else None
    codigo = m.group(2) if m else None
    nom = re.search(r'Cuenta:\s*([^\n(]+)', hoja.name)
    rango = re.search(r'(\d{2}/\d{2}/\d{4})\s+AL\s+(\d{2}/\d{2}/\d{4})', hoja.name)
    asientos, seq = [], 0
    for r in range(hoja.nrows):
        v = [hoja.cell_value(r, c) for c in range(min(hoja.ncols, 8))]
        if not isinstance(v[1], float) or str(v[0]).startswith('Saldo'):
            continue
        seq += 1
        asientos.append(AsientoMayor(
            id=f"{letra or 'E'}#{seq}", hoja=letra or 'E',
            asiento=int(v[0]) if isinstance(v[0], float) else str(v[0]),
            fecha=_fecha(v[1]), referencia=str(v[2] or '').strip(),
            comentario=str(v[3] or '').strip(),
            debe=round(float(v[4] or 0.0), 2),
            haber=round(float(v[5] or 0.0), 2)))
    return {"tipo": "fbs", "hoja": letra, "codigo_fbs": codigo,
            "codigos": {letra or "E": codigo} if codigo else {},
            "nombre_fbs": (nom.group(1).strip() if nom else hoja.name[:60]),
            "desde": _fecha(rango.group(1)) if rango else None,
            "hasta": _fecha(rango.group(2)) if rango else None,
            "asientos": asientos}


def _parse_fbs_xlsx(nombre, wb):
    """FBS exportado a .xlsx: puede traer las hojas E y O en el mismo archivo
    ("Hoja E"/"Hoja O") y sin código de cuenta; el identificador que queda para
    el mapeo es el nombre interno ("Cuenta: CIT - BANCO SANTANDER...")."""
    asientos, hojas = [], []
    codigos = {}
    nombre_fbs = None
    for ws in wb.worksheets:
        filas = list(ws.iter_rows(values_only=True))
        texto0 = ' '.join(str(v) for f in filas[:8] for v in f if v)
        if 'Consulta del Mayor' not in texto0:
            continue
        m = FBS_HOJA_RE.search(ws.title) or FBS_HOJA_RE.search(texto0)
        letra = m.group(1) if m else None
        if letra is None:
            t = re.search(r'\b(E|O)\b', ws.title.upper())
            letra = t.group(1) if t else ('O' if not hojas else 'E')
        if m:
            codigos[letra] = m.group(2)
        hojas.append(letra)
        for f in filas[:8]:
            c0 = str(f[0] or '')
            if c0.startswith('Cuenta:'):
                crudo = c0.split('Cuenta:', 1)[1].strip()
                nombre_fbs = re.split(r'\s*\((?:E|O)\)', crudo)[0].strip()[:80]
        seq = 0
        for f in filas:
            fecha = _fecha(f[1])
            if fecha is None or str(f[0] or '').startswith('Saldo') or f[0] is None:
                continue
            seq += 1
            asientos.append(AsientoMayor(
                id=f"{letra}#{seq}", hoja=letra,
                asiento=str(f[0]).strip(), fecha=fecha,
                referencia=str(f[2] or '').strip(),
                comentario=str(f[3] or '').strip(),
                debe=round(float(str(f[4] or 0).replace(',', '') or 0), 2),
                haber=round(float(str(f[5] or 0).replace(',', '') or 0), 2)))
    return {"tipo": "fbs", "hoja": ''.join(sorted(set(hojas))),
            "codigo_fbs": next(iter(codigos.values()), None), "codigos": codigos,
            "nombre_fbs": nombre_fbs, "asientos": asientos}


# --- identificación ---------------------------------------------------------

def identificar(nombre: str, data: bytes) -> dict:
    """Detecta y parsea un archivo. Siempre devuelve un dict con 'tipo'
    ('extracto' | 'fbs' | 'desconocido') y, según el caso, movimientos o
    asientos, banco, cuenta (dígitos), moneda, hoja/código FBS, rango."""
    try:
        if data[:4] == b'\xd0\xcf\x11\xe0':          # OLE2: xls binario real
            wb = xlrd.open_workbook(file_contents=data)
            nombre_hoja = wb.sheets()[0].name
            texto0 = ' '.join(str(wb.sheets()[0].cell_value(r, c))
                              for r in range(min(wb.sheets()[0].nrows, 8))
                              for c in range(min(wb.sheets()[0].ncols, 4)))
            if FBS_HOJA_RE.search(nombre_hoja) or 'Consulta del Mayor' in texto0:
                out = _parse_fbs(nombre, wb)
            elif any('CUENTA CORRIENTE BANCARIA' in str(wb.sheets()[0].cell_value(r, 0))
                     for r in range(min(wb.sheets()[0].nrows, 6))) or 'Causal' in texto0:
                out = {"tipo": "extracto", **_parse_macro(nombre, wb)}
            else:
                out = {"tipo": "extracto", **_parse_bbva(nombre, wb)}
        elif data[:2] == b'PK':                      # xlsx: FBS, Galicia o Santander
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True,
                                        data_only=True)
            celdas0 = ' '.join(str(v) for ws in wb.worksheets
                               for f in list(ws.iter_rows(values_only=True))[:8]
                               for v in f if v)
            if 'Consulta del Mayor' in celdas0:
                out = _parse_fbs_xlsx(nombre, wb)
            elif 'Grupo de Conceptos' in celdas0 or 'Movimientos' in wb.sheetnames:
                out = {"tipo": "extracto", **_parse_galicia(nombre, data)}
            elif 'Cod. Operativo' in celdas0:
                out = {"tipo": "extracto", **_parse_rio_xlsx(nombre, wb)}
            else:
                return {"tipo": "desconocido", "error":
                        "Excel no reconocido: no parece Galicia, Santander ni FBS"}
        elif _es_rio(data):
            out = {"tipo": "extracto", **_parse_rio(nombre, data)}
        elif data[:6].decode('latin-1', errors='replace').startswith('Cuenta'):
            out = {"tipo": "extracto", **_parse_ciudad(nombre, data)}
        else:
            return {"tipo": "desconocido", "error": "Formato no reconocido"}
    except Exception as exc:  # noqa: BLE001 — el error se muestra por archivo
        return {"tipo": "desconocido", "error": str(exc)}

    items = out.get("movimientos") or out.get("asientos") or []
    fechas = [x.fecha for x in items if x.fecha]
    out.setdefault("desde", min(fechas) if fechas else None)
    out.setdefault("hasta", max(fechas) if fechas else None)
    out["cantidad"] = len(items)
    out["archivo"] = nombre
    return out
