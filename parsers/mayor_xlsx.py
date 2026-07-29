# -*- coding: utf-8 -*-
"""Parser del libro mayor exportado del sistema contable (FES) en Excel.

Espera un libro con hojas cuyo nombre contenga "E" (cuenta extracto/confirmados)
y "O" (cuenta operativa/transitoria). Cada hoja tiene un encabezado
'Asiento | Fecha | Referencia | Comentario | Debe | Haber | Saldo | Centro de Costo'.
"""
import io
import re
from dataclasses import dataclass
from datetime import date, datetime

import openpyxl


@dataclass
class AsientoMayor:
    id: str
    hoja: str           # 'E' u 'O'
    asiento: int | str
    fecha: date | None
    referencia: str
    comentario: str
    debe: float
    haber: float

    @property
    def importe(self) -> float:
        return self.debe if self.debe else self.haber

    @property
    def lado(self) -> str:
        return "debe" if self.debe else "haber"

    def to_dict(self):
        return {
            "id": self.id,
            "hoja": self.hoja,
            "asiento": self.asiento,
            "fecha": self.fecha.isoformat() if self.fecha else None,
            "referencia": self.referencia,
            "comentario": self.comentario,
            "debe": self.debe,
            "haber": self.haber,
        }


def _clasificar_hoja(nombre: str) -> str | None:
    """Determina si la hoja es la cuenta E o la O por su nombre."""
    limpio = re.sub(r'[^A-Z]', ' ', nombre.upper()).split()
    if not limpio:
        return None
    # busca una 'E' u 'O' aislada al final del nombre (p.ej. "Santander E")
    for token in reversed(limpio):
        if token == 'E':
            return 'E'
        if token == 'O':
            return 'O'
    return None


def parse_mayor(source) -> dict:
    """Parsea el Excel del mayor. `source` puede ser una ruta o bytes.

    Devuelve {'E': {'saldo_inicial': float, 'asientos': [...]},
              'O': {...}, 'hojas': [nombres]}
    """
    if isinstance(source, (bytes, bytearray)):
        source = io.BytesIO(source)
    wb = openpyxl.load_workbook(source, read_only=True, data_only=True)

    resultado = {"hojas": wb.sheetnames}
    for nombre in wb.sheetnames:
        clave = _clasificar_hoja(nombre)
        if clave is None or clave in resultado:
            continue
        ws = wb[nombre]
        rows = list(ws.iter_rows(values_only=True))
        hdr = None
        saldo_inicial = None
        for i, r in enumerate(rows):
            if r and any(c == 'Saldo Inicial' for c in r if isinstance(c, str)):
                nums = [c for c in r if isinstance(c, (int, float))]
                if nums:
                    saldo_inicial = float(nums[0])
            if r and r[0] == 'Asiento':
                hdr = i
                break
        if hdr is None:
            continue
        asientos = []
        seq = 0
        for r in rows[hdr + 1:]:
            if not r or not isinstance(r[1], datetime):
                continue
            seq += 1
            asientos.append(AsientoMayor(
                id=f"{clave}#{seq}",
                hoja=clave,
                asiento=r[0],
                fecha=r[1].date(),
                referencia=str(r[2] or ''),
                comentario=str(r[3] or ''),
                debe=round(float(r[4] or 0.0), 2),
                haber=round(float(r[5] or 0.0), 2),
            ))
        resultado[clave] = {"nombre_hoja": nombre, "saldo_inicial": saldo_inicial,
                            "asientos": asientos}
    return resultado
