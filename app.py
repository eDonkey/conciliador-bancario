# -*- coding: utf-8 -*-
"""Conciliador bancario — servidor web.

Uso:
    python -m uvicorn app:app --port 8765
o simplemente:
    python app.py
"""
import io
import json
import os
import re
import sys
import tempfile
import threading
import time
import uuid
from datetime import date

from fastapi import Body, FastAPI, File, Form, UploadFile
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from parsers.santander_pdf import parse_extracto
from parsers.mayor_xlsx import parse_mayor
from engine.matcher import conciliar
from engine import ai_assist
from engine import reglas as reglas_mod
from engine import equivalencias as eq_mod
from engine import gastos_conf
from engine import cuentas as cuentas_mod
from engine import analisis as analisis_mod
from parsers import diarios

app = FastAPI(title="Conciliador bancario")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATOS_DIR = os.path.join(BASE_DIR, "datos")
os.makedirs(DATOS_DIR, exist_ok=True)
RESULTADOS = {}  # job_id -> resultado serializado
PROGRESO = {}    # job_id -> {estado, fase, porcentaje, eta_seg, ...}
STAGING = {}     # staging_id -> archivos parseados del modo diario

LISTAS_BANCO = ["banco_sin_contabilizar", "gastos_bancarios"]
LISTAS_MAYOR = ["e_sin_banco", "o_pendientes_sin_banco"]


def _guardar(job_id):
    with open(os.path.join(DATOS_DIR, f"{job_id}.json"), "w", encoding="utf-8") as f:
        json.dump(RESULTADOS[job_id], f, ensure_ascii=False)


def _obtener(job_id):
    if job_id not in RESULTADOS:
        ruta = os.path.join(DATOS_DIR, f"{job_id}.json")
        if os.path.exists(ruta):
            with open(ruta, encoding="utf-8") as f:
                RESULTADOS[job_id] = json.load(f)
    return RESULTADOS.get(job_id)


def _recalcular_resumen(datos):
    r = datos["resumen"]
    imp_b = lambda x: x["credito"] or x["debito"]
    imp_m = lambda x: x["debe"] or x["haber"]
    for lista in LISTAS_BANCO:
        r[lista] = {"cantidad": len(datos[lista]),
                    "importe": round(sum(imp_b(x) for x in datos[lista]), 2)}
    for lista in LISTAS_MAYOR:
        r[lista] = {"cantidad": len(datos[lista]),
                    "importe": round(sum(imp_m(x) for x in datos[lista]), 2)}
    r["conciliados_manual"] = len(datos.get("conciliados_manual", []))
    movs_manual = sum(len(m["banco"]) for m in datos.get("conciliados_manual", []))
    conciliados = r["conciliados_e"] + r["en_o_pendientes_confirmar"] + movs_manual
    r["porcentaje_conciliado"] = round(
        100.0 * conciliados / max(1, r["movimientos_banco"]), 1)
    gastos_expl = sum(p.get("movimientos", 0) for p in datos.get("nota_debito", [])
                      if p.get("total_mayor") is not None)
    r["gastos_explicados"] = gastos_expl
    r["porcentaje_explicado"] = round(
        min(100.0, 100.0 * (conciliados + gastos_expl) / max(1, r["movimientos_banco"])), 1)


def _serializar(resultado, extractos, ia_sugerencias, ia_estado):
    def match_dict(m):
        return {"banco": m["banco"].to_dict(), "asiento": m["asiento"].to_dict(),
                "metodo": m["metodo"]}

    return {
        "extractos": [
            {k: (v.isoformat() if hasattr(v, "isoformat") else v)
             for k, v in e.items() if k != "movimientos"}
            for e in extractos
        ],
        "resumen": resultado["resumen"],
        "o_cancelados": resultado["o_cancelados"],
        "o_pendientes_total": resultado["o_pendientes_total"],
        "conciliados_e": [match_dict(m) for m in resultado["matches_e"]],
        "conciliados_o": [dict(match_dict(m), confirmado=False)
                          for m in resultado["matches_o"]],
        "banco_sin_contabilizar": [m.to_dict() for m in resultado["banco_sin_contabilizar"]],
        "gastos_bancarios": [m.to_dict() for m in resultado["gastos_bancarios"]],
        "e_sin_banco": [a.to_dict() for a in resultado["e_sin_banco"]],
        "o_pendientes_sin_banco": [a.to_dict() for a in resultado["o_pendientes_sin_banco"]],
        "nota_debito": [
            {**p, "asientos_mayor": [a.to_dict() for a in p["asientos_mayor"]]}
            for p in resultado["nota_debito"]
        ],
        "contraasientos": [
            {"debe": p["debe"].to_dict(), "haber": p["haber"].to_dict()}
            for p in resultado["contraasientos"]
        ],
        "excluidos_mayor": [],
        "ia": {"estado": ia_estado, "sugerencias": ia_sugerencias},
    }


@app.post("/api/conciliar")
async def api_conciliar(
    extractos: list[UploadFile] = File(...),
    mayor: UploadFile = File(...),
    usar_ia: str = Form("no"),
):
    """Recibe los archivos, arranca el procesamiento en segundo plano y
    devuelve el job_id de inmediato. El avance se consulta en /api/progreso."""
    archivos = [(up.filename, await up.read()) for up in extractos]
    data_mayor = await mayor.read()

    job_id = uuid.uuid4().hex[:12]
    total_mb = (sum(len(b) for _, b in archivos) + len(data_mayor)) / 1_000_000
    est_base = 10 + total_mb * 14   # parseo + cruce, calibrado con datos reales
    PROGRESO[job_id] = {"estado": "procesando", "fase": "Preparando el análisis",
                        "porcentaje": 2, "eta_seg": round(est_base),
                        "inicio": time.time()}
    threading.Thread(target=_procesar_job,
                     args=(job_id, archivos, data_mayor, usar_ia, est_base),
                     daemon=True).start()
    return {"job_id": job_id, "estado": "procesando"}


def _procesar_job(job_id, archivos, data_mayor, usar_ia, est_base):
    inicio = PROGRESO[job_id]["inicio"]

    def prog(fase, pct, eta=None):
        if eta is None:
            eta = max(3, est_base - (time.time() - inicio))
        PROGRESO[job_id] = {"estado": "procesando", "fase": fase,
                            "porcentaje": round(pct), "eta_seg": round(eta),
                            "inicio": inicio}

    try:
        # --- parsear extractos -------------------------------------------
        parseados = []
        n = len(archivos)
        for i, (nombre, data) in enumerate(archivos):
            prog(f"Leyendo extracto {i + 1} de {n}: {nombre}", 5 + 30 * i / n)
            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
                tmp.write(data)
                ruta = tmp.name
            try:
                parseados.append(parse_extracto(ruta, nombre))
            finally:
                os.unlink(ruta)
        parseados.sort(key=lambda e: (e["desde"] or e["hasta"] or ""))
        movs = [m for e in parseados for m in e["movimientos"]]

        # --- parsear mayor -----------------------------------------------
        prog("Leyendo el libro mayor", 38)
        mayor_parsed = parse_mayor(data_mayor)
        if "E" not in mayor_parsed or "O" not in mayor_parsed:
            PROGRESO[job_id] = {"estado": "error", "mensaje":
                "El Excel debe tener una hoja de la cuenta E y una de la cuenta O "
                f"(hojas encontradas: {mayor_parsed['hojas']})"}
            return

        # --- conciliar ----------------------------------------------------
        prog("Cruzando el extracto contra el mayor", 46)
        reglas = reglas_mod.cargar()
        equivalencias = eq_mod.cargar()
        terminos_gasto = gastos_conf.cargar()
        resultado = conciliar(movs, mayor_parsed["E"]["asientos"],
                              mayor_parsed["O"]["asientos"], reglas_aprendidas=reglas,
                              equivalencias=equivalencias,
                              terminos_gasto=terminos_gasto)

        # --- IA (real o simulada) sobre los residuales --------------------
        ia_sugerencias, ia_estado = [], "desactivada"
        if usar_ia in ("si", "simular"):
            residual_mayor = (resultado["e_sin_banco"]
                              + resultado["o_pendientes_sin_banco"])
            # los gastos bancarios no van a la IA: ya los explica la ND mensual
            residual_banco = resultado["banco_sin_contabilizar"]
            seg_tanda = 4 if usar_ia == "simular" else 45

            def prog_ia(tanda, total):
                prog(f"Análisis con IA — tanda {tanda} de {total}",
                     58 + 38 * (tanda - 1) / total,
                     eta=(total - tanda + 1) * seg_tanda)

            if usar_ia == "simular":
                ia_sugerencias = ai_assist.simular_sugerencias(
                    residual_banco, residual_mayor, progreso=prog_ia,
                    glosario=equivalencias)
                ia_estado = "simulada"
            elif not ai_assist.disponible():
                ia_estado = "sin_credenciales"
            else:
                try:
                    ia_sugerencias = ai_assist.sugerir_matches(
                        residual_banco, residual_mayor, progreso=prog_ia,
                        glosario=equivalencias)
                    ia_estado = "ok"
                except Exception as exc:  # noqa: BLE001 — mostrar el error
                    ia_estado = f"error: {exc}"

        prog("Preparando los resultados", 96, eta=4)
        salida = _serializar(resultado, parseados, ia_sugerencias, ia_estado)
        # auditoría: qué modo de IA llegó en el pedido (para diagnosticar
        # corridas en simulación no intencionales)
        salida["ia"]["modo_pedido"] = usar_ia
        salida["conciliados_manual"] = []
        salida["resumen"]["conciliados_manual"] = 0
        salida["resumen"]["reglas_disponibles"] = len(reglas)
        salida["job_id"] = job_id
        analisis_mod.anotar_residuales(salida)
        RESULTADOS[job_id] = salida
        _guardar(job_id)
        PROGRESO[job_id] = {"estado": "listo", "porcentaje": 100, "eta_seg": 0,
                            "fase": "Conciliación terminada", "inicio": inicio}
    except Exception as exc:  # noqa: BLE001 — que el error llegue a la UI
        PROGRESO[job_id] = {"estado": "error", "mensaje": str(exc)}


@app.get("/api/diagnostico")
def api_diagnostico():
    """Qué ve el servidor del entorno de IA, sin exponer secretos. Sirve para
    verificar que la clave llegó al proceso (p. ej. tras un deploy)."""
    clave = os.environ.get("ANTHROPIC_API_KEY") or ""
    token = os.environ.get("ANTHROPIC_AUTH_TOKEN") or ""
    return {
        "ia_disponible": ai_assist.disponible(),
        "ANTHROPIC_API_KEY": {
            "presente": bool(clave.strip()),
            "longitud": len(clave),
            "formato_esperado": clave.strip().startswith("sk-ant-"),
            "espacios_al_borde": clave != clave.strip(),
            "entre_comillas": clave.strip()[:1] in ('"', "'") if clave.strip() else False,
        },
        "ANTHROPIC_AUTH_TOKEN_presente": bool(token.strip()),
        "archivo_clave_local": os.path.exists(ai_assist.RUTA_CLAVE),
        # True si datos/ está montado como volumen persistente (Railway):
        # sin esto, todo el aprendizaje se pierde en cada deploy
        "datos_es_volumen": os.path.ismount(DATOS_DIR),
        # repr() revela caracteres invisibles en el nombre (espacios al final)
        "variables_con_nombre_parecido": sorted(
            repr(k) for k in os.environ
            if "ANTHROPIC" in k.upper() or "API_KEY" in k.upper()),
    }


@app.post("/api/analizar/{job_id}")
def api_analizar(job_id: str, cuerpo: dict = Body(...)):
    """Análisis con IA de un asiento pendiente: por qué quedó sin banco.
    Cachea el resultado en el job para no repetir la llamada."""
    datos = _obtener(job_id)
    if not datos:
        return JSONResponse(status_code=404, content={"error": "Resultado no encontrado"})
    id_mayor = cuerpo.get("id_mayor")
    asiento = next((x for lista in LISTAS_MAYOR for x in datos[lista]
                    if x["id"] == id_mayor), None)
    if not asiento:
        return JSONResponse(status_code=404, content={"error": "Asiento no encontrado"})

    cache = datos.setdefault("analisis", {})
    if id_mayor in cache and not cuerpo.get("rehacer"):
        return {"analisis": cache[id_mayor]}

    aprendidos = analisis_mod.cargar()

    # si la firma ya tiene un veredicto del usuario, no se gasta IA: se
    # devuelve directamente lo aprendido (regla o corrección)
    firma = reglas_mod.firma_mayor(asiento.get("referencia"), asiento.get("comentario"))
    previo = analisis_mod.buscar(firma, aprendidos)
    if previo and not cuerpo.get("rehacer"):
        cache[id_mayor] = {
            "texto": previo.get("correccion") or previo.get("explicacion", ""),
            "fecha": date.today().isoformat(),
            "veredicto": previo["veredicto"],
            "correccion": previo.get("correccion", ""),
            "origen": "aprendizaje",
        }
        _guardar(job_id)
        return {"analisis": cache[id_mayor]}

    correcciones = [a for a in aprendidos if a["veredicto"] == "corregido"][:8]
    resultado = analisis_mod.analizar_asiento(
        asiento, datos.get("banco_sin_contabilizar", []), correcciones,
        cuenta=(datos.get("cuenta") or {}).get("etiqueta", ""),
        simular=bool(cuerpo.get("simular")))

    cache[id_mayor] = {"texto": resultado["texto"], "fecha": date.today().isoformat(),
                       "veredicto": None, "correccion": "",
                       "origen": resultado["origen"]}
    _guardar(job_id)
    return {"analisis": cache[id_mayor]}


@app.post("/api/analizar/{job_id}/veredicto")
def api_analizar_veredicto(job_id: str, cuerpo: dict = Body(...)):
    """Confirmación del usuario sobre un análisis: Sí -> se convierte en
    regla; No + explicación -> el sistema aprende la corrección."""
    datos = _obtener(job_id)
    if not datos:
        return JSONResponse(status_code=404, content={"error": "Resultado no encontrado"})
    id_mayor = cuerpo.get("id_mayor")
    entrada = (datos.get("analisis") or {}).get(id_mayor)
    if not entrada:
        return JSONResponse(status_code=409, content={
            "error": "Primero hay que correr el análisis de este asiento"})
    ok = bool(cuerpo.get("ok"))
    correccion = (cuerpo.get("correccion") or "").strip()
    if not ok and not correccion:
        return JSONResponse(status_code=422, content={
            "error": "Contanos por qué el análisis no está bien, así el sistema aprende"})

    asiento = next((x for lista in LISTAS_MAYOR for x in datos[lista]
                    if x["id"] == id_mayor), None)
    if asiento:
        firma = reglas_mod.firma_mayor(asiento.get("referencia"), asiento.get("comentario"))
        analisis_mod.aprender(firma, entrada["texto"], ok, correccion)
        analisis_mod.anotar_residuales(datos)
    entrada["veredicto"] = "ok" if ok else "corregido"
    entrada["correccion"] = correccion
    _guardar(job_id)
    return {"analisis": entrada}


ARCHIVO_MIGRABLE_RE = re.compile(r'^[\w.-]+\.json$')


@app.get("/api/migracion/exportar")
def api_migracion_exportar():
    """Exporta todos los datos del servicio (aprendizajes, resultados,
    tableros, cuentas) como un solo JSON. Nunca incluye credenciales."""
    bundle = {}
    for nombre in sorted(os.listdir(DATOS_DIR)):
        if not ARCHIVO_MIGRABLE_RE.match(nombre):
            continue
        try:
            with open(os.path.join(DATOS_DIR, nombre), encoding="utf-8") as f:
                bundle[nombre] = json.load(f)
        except (json.JSONDecodeError, OSError):
            continue
    return {"version": 1, "exportado": date.today().isoformat(),
            "cantidad": len(bundle), "archivos": bundle}


@app.post("/api/migracion/importar")
def api_migracion_importar(cuerpo: dict = Body(...)):
    """Importa un bundle exportado desde otro servicio. Como protección,
    solo funciona sobre un servicio SIN conciliaciones diarias previas
    (salvo {"forzar": true}), para no pisar datos productivos por error."""
    hay_grupos = any(n.startswith("grupo_") for n in os.listdir(DATOS_DIR))
    if hay_grupos and not cuerpo.get("forzar"):
        return JSONResponse(status_code=409, content={
            "error": "Este servicio ya tiene conciliaciones cargadas. "
                     "Para pisarlas igual hay que mandar forzar=true."})
    archivos = cuerpo.get("archivos") or {}
    escritos, ignorados = 0, []
    for nombre, contenido in archivos.items():
        if not ARCHIVO_MIGRABLE_RE.match(nombre):
            ignorados.append(nombre)
            continue
        with open(os.path.join(DATOS_DIR, nombre), "w", encoding="utf-8") as f:
            json.dump(contenido, f, ensure_ascii=False)
        escritos += 1
    RESULTADOS.clear()   # invalidar el caché en memoria: releer del disco
    return {"ok": True, "archivos_importados": escritos, "ignorados": ignorados,
            "datos_es_volumen": os.path.ismount(DATOS_DIR)}


@app.get("/api/equivalencias")
def api_equivalencias():
    """Diccionario de equivalencias de términos extracto ↔ sistema."""
    return {"equivalencias": eq_mod.cargar()}


@app.post("/api/equivalencias")
def api_equivalencias_agregar(cuerpo: dict = Body(...)):
    pares, error = eq_mod.agregar(cuerpo.get("extracto"), cuerpo.get("sistema"))
    if error:
        return JSONResponse(status_code=422, content={"error": error})
    return {"equivalencias": pares}


@app.delete("/api/equivalencias/{eq_id}")
def api_equivalencias_eliminar(eq_id: str):
    return {"equivalencias": eq_mod.eliminar(eq_id)}


@app.get("/api/cuentas")
def api_cuentas():
    """Cuentas bancarias del grupo con su etiqueta y mapeo FBS."""
    cuentas = cuentas_mod.cargar()
    return {"cuentas": [{**c, "etiqueta": cuentas_mod.etiqueta(c)} for c in cuentas],
            "bancos": cuentas_mod.BANCOS}


@app.post("/api/diario/identificar")
async def api_diario_identificar(archivos: list[UploadFile] = File(...)):
    """Modo diario, paso 1: detecta qué es cada archivo (banco/cuenta/FBS) y
    deja lo parseado en memoria para el paso de conciliación."""
    cuentas = cuentas_mod.cargar()
    staging_id = uuid.uuid4().hex[:10]
    parseados, resumen = {}, []
    for up in archivos:
        data = await up.read()
        info = diarios.identificar(up.filename, data)
        parseados[up.filename] = info
        cuenta = None
        if info["tipo"] == "extracto":
            cuenta = cuentas_mod.buscar_por_numero(
                cuentas, info.get("banco"), info.get("cuenta"), info.get("moneda"))
        elif info["tipo"] == "fbs":
            idents = list((info.get("codigos") or {}).values()) + [info.get("nombre_fbs")]
            cuenta = next((c for c in (cuentas_mod.buscar_por_fbs(cuentas, i)
                                       for i in idents if i) if c), None)
            if cuenta is None:  # el nombre interno a veces trae el nro real
                cuenta = cuentas_mod.buscar_por_nombre_fbs(cuentas, info.get("nombre_fbs"))
        resumen.append({
            "archivo": up.filename, "tipo": info["tipo"],
            "banco": info.get("banco"), "moneda": info.get("moneda"),
            "cuenta_detectada": info.get("cuenta"),
            "hoja": info.get("hoja"), "codigo_fbs": info.get("codigo_fbs"),
            "nombre_fbs": info.get("nombre_fbs"),
            "cantidad": info.get("cantidad", 0),
            "desde": info["desde"].isoformat() if info.get("desde") else None,
            "hasta": info["hasta"].isoformat() if info.get("hasta") else None,
            "cuenta_id": cuenta["id"] if cuenta else None,
            "error": info.get("error"),
        })
    STAGING[staging_id] = {"archivos": parseados, "resumen": resumen}
    return {"staging_id": staging_id, "archivos": resumen}


RUTA_ARRASTRE = os.path.join(DATOS_DIR, "arrastre_diario.json")


def _cargar_arrastre() -> dict:
    if os.path.exists(RUTA_ARRASTRE):
        try:
            with open(RUTA_ARRASTRE, encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, OSError):
            return {}
    return {}


def _guardar_arrastre(d: dict):
    with open(RUTA_ARRASTRE, "w", encoding="utf-8") as f:
        json.dump(d, f, ensure_ascii=False, indent=1)


def _fecha_iso(s):
    from datetime import date as _d
    return _d.fromisoformat(s) if s else None


def _reconstruir_mov(d, seq):
    from parsers.santander_pdf import MovimientoBanco
    archivo = d.get("archivo") or ""
    if not archivo.startswith("(pendiente"):
        archivo = f"(pendiente de días anteriores) {archivo}"
    return MovimientoBanco(
        id=f"P#{seq}", fecha=_fecha_iso(d.get("fecha")),
        comprobante=d.get("comprobante"), descripcion=d.get("descripcion") or "",
        detalle=d.get("detalle") or "", debito=d.get("debito") or 0.0,
        credito=d.get("credito") or 0.0, saldo=d.get("saldo") or 0.0,
        archivo=archivo, pagina=0)


def _reconstruir_asiento(d, seq):
    from parsers.mayor_xlsx import AsientoMayor
    return AsientoMayor(
        id=f"P{d.get('hoja', 'E')}#{seq}", hoja=d.get("hoja", "E"),
        asiento=d.get("asiento"), fecha=_fecha_iso(d.get("fecha")),
        referencia=d.get("referencia") or "", comentario=d.get("comentario") or "",
        debe=d.get("debe") or 0.0, haber=d.get("haber") or 0.0)


def _aplicar_arrastre(cid, movs, g, arrastre):
    """Suma al cruce los residuales de la conciliación anterior de la cuenta:
    movimientos del banco sin contabilizar y asientos FBS sin banco, salteando
    los que ya vienen en los archivos de hoy. Devuelve (n_movs, n_asientos)."""
    prev = _obtener(arrastre.get(cid) or "")
    if not prev:
        return 0, 0
    seq = 0
    claves_m = {(m.fecha, m.descripcion, round(m.credito - m.debito, 2), m.comprobante)
                for m in movs}
    n_movs = 0
    for d in prev.get("banco_sin_contabilizar", []):
        clave = (_fecha_iso(d.get("fecha")), d.get("descripcion") or "",
                 round((d.get("credito") or 0) - (d.get("debito") or 0), 2),
                 d.get("comprobante"))
        if clave in claves_m:
            continue
        seq += 1
        movs.append(_reconstruir_mov(d, seq))
        n_movs += 1
    claves_a = {(a.hoja, str(a.asiento), a.fecha, round(a.debe, 2), round(a.haber, 2))
                for a in g["e"] + g["o"]}
    n_asientos = 0
    for lista, dest in (("e_sin_banco", g["e"]), ("o_pendientes_sin_banco", g["o"])):
        for d in prev.get(lista, []):
            clave = (d.get("hoja", "E"), str(d.get("asiento")), _fecha_iso(d.get("fecha")),
                     round(d.get("debe") or 0, 2), round(d.get("haber") or 0, 2))
            if clave in claves_a:
                continue
            seq += 1
            dest.append(_reconstruir_asiento(d, seq))
            n_asientos += 1
    return n_movs, n_asientos


# --- memoria de conciliados: los archivos acumulativos no duplican ----------
# El cliente sube el extracto del día pero el mayor FBS acumulado (o a fin de
# mes, todo junto). Sin esto, lo ya conciliado en corridas anteriores
# reaparecía como "sin banco"/"sin asiento". La memoria guarda, por cuenta,
# cuántas veces se consumió cada movimiento/asiento, y las corridas nuevas
# omiten esas repeticiones.
RUTA_MEMORIA = os.path.join(DATOS_DIR, "memoria_conciliados.json")


def _cargar_memoria() -> dict:
    if os.path.exists(RUTA_MEMORIA):
        try:
            with open(RUTA_MEMORIA, encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, OSError):
            return {}
    return {}


def _guardar_memoria(m: dict):
    with open(RUTA_MEMORIA, "w", encoding="utf-8") as f:
        json.dump(m, f, ensure_ascii=False)


def _clave_mov_dict(d) -> str:
    neto = (d.get("credito") or 0) - (d.get("debito") or 0)
    return f'{d.get("fecha") or ""}|{d.get("descripcion") or ""}|{neto:.2f}|{d.get("comprobante") or ""}'


def _clave_mov_obj(m) -> str:
    return (f'{m.fecha.isoformat() if m.fecha else ""}|{m.descripcion}|'
            f'{m.credito - m.debito:.2f}|{m.comprobante or ""}')


def _clave_asiento_dict(d) -> str:
    return (f'{d.get("hoja") or ""}|{d.get("asiento") or ""}|{d.get("fecha") or ""}|'
            f'{d.get("debe") or 0:.2f}|{d.get("haber") or 0:.2f}')


def _clave_asiento_obj(a) -> str:
    return (f'{a.hoja}|{a.asiento}|{a.fecha.isoformat() if a.fecha else ""}|'
            f'{a.debe:.2f}|{a.haber:.2f}')


def _cosechar_consumidos(prev: dict):
    """Claves de todo lo ya conciliado/explicado en un job, en su estado
    ACTUAL (incluye lo conciliado a mano después de la corrida)."""
    movs, asientos = {}, {}

    def add(d, k):
        d[k] = d.get(k, 0) + 1

    for m in prev.get("conciliados_e", []) + prev.get("conciliados_o", []):
        add(movs, _clave_mov_dict(m["banco"]))
        add(asientos, _clave_asiento_dict(m["asiento"]))
    for m in prev.get("conciliados_manual", []):
        for b in m.get("banco", []):
            add(movs, _clave_mov_dict(b))
        for a in m.get("mayor", []):
            add(asientos, _clave_asiento_dict(a))
    for b in prev.get("gastos_bancarios", []):
        add(movs, _clave_mov_dict(b))
    for p in prev.get("nota_debito", []):
        for a in p.get("asientos_mayor", []):
            add(asientos, _clave_asiento_dict(a))
    for p in prev.get("contraasientos", []):
        add(asientos, _clave_asiento_dict(p["debe"]))
        add(asientos, _clave_asiento_dict(p["haber"]))
    for e in prev.get("excluidos_mayor", []):
        for a in e.get("mayor", []):
            add(asientos, _clave_asiento_dict(a))
    return movs, asientos


def _procesar_confirmaciones(prev, g):
    """Cierra el ciclo O -> E entre corridas diarias.

    Los movimientos que en la corrida anterior quedaron "conciliados contra O,
    pendientes de confirmar" se arrastran esperando su confirmación: si en los
    archivos de hoy aparece su asiento en la cuenta E (ya se cargaron en FBS),
    pasan DIRECTO a Conciliados (E) y se cancela la contrapartida que la
    confirmación generó en O. Si todavía no se confirmaron, siguen figurando
    como pendientes de confirmar.

    Devuelve (confirmados_e, pendientes_o_arrastrados, contrapartidas_canceladas)."""
    from engine import matcher as matcher_mod

    esperando = list((prev or {}).get("conciliados_o") or [])
    if not esperando or not g["e"]:
        return [], esperando, 0

    movs = []
    for i, ent in enumerate(esperando):
        m = _reconstruir_mov(dict(ent["banco"]), i)
        m.id = f"C#{i}"
        m.archivo = ent["banco"].get("archivo") or ""
        movs.append(m)
    matches, _, e_restante = matcher_mod._match_pases(movs, g["e"])
    g["e"][:] = e_restante

    confirmados, idx_conf, canceladas = [], set(), 0
    for mt in matches:
        i = int(mt["banco"].id[2:])
        idx_conf.add(i)
        ent = esperando[i]
        confirmados.append({
            "banco": ent["banco"], "asiento": mt["asiento"].to_dict(),
            "metodo": "confirmado en E (venía conciliado contra O)",
        })
        # la confirmación genera una contrapartida en O: cancelarla del pool
        aso = ent.get("asiento") or {}
        clave = (matcher_mod._clave_rm(aso.get("comentario") or "")
                 or (aso.get("referencia") or "").strip().upper(),
                 round((aso.get("debe") or 0) or (aso.get("haber") or 0), 2))
        lado_original = "debe" if (aso.get("debe") or 0) else "haber"
        for j, a in enumerate(g["o"]):
            clave_a = (matcher_mod._clave_rm(a.comentario) or a.referencia.strip().upper(),
                       round(a.importe, 2))
            if clave_a == clave and a.lado != lado_original:
                del g["o"][j]
                canceladas += 1
                break

    pendientes = []
    for i, ent in enumerate(esperando):
        if i in idx_conf:
            continue
        metodo = ent.get("metodo") or ""
        if "arrastrado" not in metodo:
            ent = {**ent, "metodo": (metodo + " · arrastrado").strip(" ·")}
        pendientes.append(ent)
    return confirmados, pendientes, canceladas


def _resumen_mini(r):
    return {
        "movimientos_banco": r["movimientos_banco"],
        "conciliados": r["conciliados_e"] + r["en_o_pendientes_confirmar"]
                       + r.get("conciliados_manual", 0),
        "porcentaje_explicado": r.get("porcentaje_explicado", r["porcentaje_conciliado"]),
        "banco_sin_contabilizar": r["banco_sin_contabilizar"]["cantidad"],
        "gastos_bancarios": r["gastos_bancarios"]["cantidad"],
        "mayor_sin_banco": r["e_sin_banco"]["cantidad"] + r["o_pendientes_sin_banco"]["cantidad"],
        "arrastrados": r.get("arrastrados", 0),
        "arrastrados_movs": r.get("arrastrados_movs", 0),
        "arrastrados_asientos": r.get("arrastrados_asientos", 0),
        "omitidos": r.get("omitidos", 0),
    }


@app.post("/api/diario/conciliar/{staging_id}")
def api_diario_conciliar(staging_id: str, cuerpo: dict = Body(...)):
    """Modo diario, paso 2: con las asignaciones archivo→cuenta confirmadas,
    corre una conciliación por cuenta (mismo motor y mismos aprendizajes que
    el modo mensual) y devuelve el tablero del día."""
    stag = STAGING.get(staging_id)
    if not stag:
        return JSONResponse(status_code=404, content={
            "error": "La identificación expiró (el servidor se reinició). Volvé a subir los archivos."})
    asignaciones = cuerpo.get("asignaciones") or {}
    cuentas = cuentas_mod.cargar()
    por_id = {c["id"]: c for c in cuentas}

    # agrupar archivos por cuenta y aprender mapeos FBS nuevos
    grupos = {}
    sin_asignar = []
    for fila in stag["resumen"]:
        nombre = fila["archivo"]
        cid = asignaciones.get(nombre, fila.get("cuenta_id"))
        if not cid or cid == "ignorar" or fila["tipo"] == "desconocido":
            sin_asignar.append({"archivo": nombre, "tipo": fila["tipo"],
                                "motivo": fila.get("error") or "sin cuenta asignada"})
            continue
        info = stag["archivos"][nombre]
        g = grupos.setdefault(cid, {"movs": [], "e": [], "o": [], "archivos": []})
        g["archivos"].append(nombre)
        if fila["tipo"] == "extracto":
            g["movs"].extend(info["movimientos"])
        else:
            # un archivo FBS puede traer la hoja E, la O, o las dos juntas
            g["e"].extend(a for a in info["asientos"] if a.hoja == "E")
            g["o"].extend(a for a in info["asientos"] if a.hoja == "O")
            cods = info.get("codigos") or {}
            if por_id.get(cid):
                for letra in set(info.get("hoja") or "E") & {"E", "O"}:
                    ident = cods.get(letra) or fila.get("nombre_fbs")
                    if ident and ident not in (por_id[cid].get("fbs_e"),
                                               por_id[cid].get("fbs_o")):
                        cuentas = cuentas_mod.mapear_fbs(cid, letra, ident,
                                                         fila.get("nombre_fbs"))
                por_id = {c["id"]: c for c in cuentas}

    reglas = reglas_mod.cargar()
    equivalencias = eq_mod.cargar()
    terminos_gasto = gastos_conf.cargar()

    arrastre = _cargar_arrastre()
    memoria = _cargar_memoria()
    tablero = []
    for cid, g in grupos.items():
        etiqueta = cuentas_mod.etiqueta(por_id[cid]) if cid in por_id else cid

        # memoria de conciliados: sumar lo consumido por la corrida anterior
        # (en su estado actual, manuales incluidos) y omitir las repeticiones
        # que traigan los archivos acumulativos de hoy
        mem = memoria.setdefault(cid, {"movs": {}, "asientos": {}, "jobs": []})
        mem.setdefault("jobs", [])
        prev_id = arrastre.get(cid)
        prev = _obtener(prev_id or "")
        if prev and prev_id not in mem["jobs"]:   # cada corrida se cosecha una sola vez
            c_movs, c_asientos = _cosechar_consumidos(prev)
            for k, n in c_movs.items():
                mem["movs"][k] = mem["movs"].get(k, 0) + n
            for k, n in c_asientos.items():
                mem["asientos"][k] = mem["asientos"].get(k, 0) + n
            mem["jobs"].append(prev_id)

        om_movs = om_asientos = 0
        usados_mem = {}
        # dedupe SOLO entre archivos distintos (p. ej. el extracto "del día" y el
        # "histórico" subidos por separado). Dentro de un mismo archivo, filas
        # idénticas son movimientos reales (comisiones repetidas) y se conservan.
        vistos, movs = {}, []
        for m in sorted(g["movs"], key=lambda x: (x.fecha or date.min, x.id)):
            k = _clave_mov_obj(m)
            if usados_mem.get(k, 0) < mem["movs"].get(k, 0):
                usados_mem[k] = usados_mem.get(k, 0) + 1
                om_movs += 1
                continue
            clave = (m.fecha, m.descripcion, round(m.credito - m.debito, 2),
                     m.saldo, m.comprobante)
            if clave in vistos and vistos[clave] != m.archivo:
                continue
            vistos.setdefault(clave, m.archivo)
            m.id = f"B#{len(movs) + 1}"
            movs.append(m)
        usados_mem = {}
        for lado in ("e", "o"):
            filtrados = []
            for a in g[lado]:
                k = _clave_asiento_obj(a)
                if usados_mem.get(k, 0) < mem["asientos"].get(k, 0):
                    usados_mem[k] = usados_mem.get(k, 0) + 1
                    om_asientos += 1
                    continue
                filtrados.append(a)
            g[lado] = filtrados
        # ciclo O -> E: lo conciliado contra O en la corrida anterior espera su
        # confirmación; si el E de hoy la trae, pasa directo a Conciliados (E)
        conf_e, carried_o, contrap_cancel = _procesar_confirmaciones(prev, g)

        # arrastre: los pendientes de la conciliación anterior de esta cuenta
        # entran al cruce de hoy (lo de ayer aparece en el FBS de hoy y viceversa)
        arr_movs, arr_asientos = _aplicar_arrastre(cid, movs, g, arrastre)
        if not movs:
            tablero.append({"cuenta_id": cid, "etiqueta": etiqueta, "estado": "sin_extracto",
                            "archivos": g["archivos"]})
            continue
        if not g["e"] and not g["o"]:
            tablero.append({"cuenta_id": cid, "etiqueta": etiqueta, "estado": "sin_fbs",
                            "archivos": g["archivos"],
                            "resumen": {"movimientos_banco": len(movs)}})
            continue
        resultado = conciliar(movs, g["e"], g["o"], reglas_aprendidas=reglas,
                              equivalencias=equivalencias, terminos_gasto=terminos_gasto)
        # período cubierto por los extractos de HOY (sin contar los arrastrados,
        # que traen fechas de días anteriores)
        fechas = [m.fecha for m in movs if m.fecha and m.id.startswith("B#")]
        extractos_meta = [{"archivo": ", ".join(g["archivos"]),
                           "desde": min(fechas) if fechas else None,
                           "hasta": max(fechas) if fechas else None}]
        job_id = uuid.uuid4().hex[:12]
        salida = _serializar(resultado, extractos_meta, [], "desactivada")
        salida["conciliados_manual"] = []
        salida["resumen"]["conciliados_manual"] = 0
        salida["resumen"]["reglas_disponibles"] = len(reglas)
        salida["job_id"] = job_id
        salida["cuenta"] = {"id": cid, "etiqueta": etiqueta}
        # ciclo O -> E: sumar confirmados y arrastrar los que siguen pendientes
        if conf_e or carried_o:
            salida["conciliados_e"].extend(conf_e)
            salida["conciliados_o"].extend(carried_o)
            r = salida["resumen"]
            r["conciliados_e"] += len(conf_e)
            r["en_o_pendientes_confirmar"] += len(carried_o)
            r["confirmados_desde_o"] = len(conf_e)
            r["contrapartidas_o_canceladas"] = contrap_cancel
            r["movimientos_banco"] += len(conf_e) + len(carried_o)
            conc = r["conciliados_e"] + r["en_o_pendientes_confirmar"]
            r["porcentaje_conciliado"] = round(
                100.0 * conc / max(1, r["movimientos_banco"]), 1)
            r["porcentaje_explicado"] = round(min(100.0,
                100.0 * (conc + r.get("gastos_explicados", 0))
                / max(1, r["movimientos_banco"])), 1)

        # de qué job venía la cadena de esta cuenta (para poder deshacer corridas)
        salida["arrastre_desde"] = arrastre.get(cid)
        if om_movs or om_asientos:
            salida["memoria_omitidos"] = {"movimientos": om_movs, "asientos": om_asientos}
            salida["resumen"]["omitidos"] = om_movs + om_asientos
        analisis_mod.anotar_residuales(salida)
        if arr_movs or arr_asientos:
            salida["arrastre"] = {"movimientos": arr_movs, "asientos": arr_asientos,
                                  "desde_job": arrastre.get(cid)}
            salida["resumen"]["arrastrados"] = arr_movs + arr_asientos
            salida["resumen"]["arrastrados_movs"] = arr_movs
            salida["resumen"]["arrastrados_asientos"] = arr_asientos
        RESULTADOS[job_id] = salida
        _guardar(job_id)
        arrastre[cid] = job_id   # la próxima conciliación arrastra desde acá
        tablero.append({"cuenta_id": cid, "etiqueta": etiqueta, "estado": "ok",
                        "job_id": job_id, "archivos": g["archivos"],
                        "desde": fechas and min(fechas).isoformat() or None,
                        "hasta": fechas and max(fechas).isoformat() or None,
                        "resumen": _resumen_mini(salida["resumen"])})

    _guardar_arrastre(arrastre)
    _guardar_memoria(memoria)
    grupo_id = uuid.uuid4().hex[:10]
    grupo = {"grupo_id": grupo_id, "procesado": date.today().isoformat(),
             "hora": time.strftime("%H:%M"),
             "cuentas": sorted(tablero, key=lambda x: x["etiqueta"]),
             "sin_asignar": sin_asignar}
    with open(os.path.join(DATOS_DIR, f"grupo_{grupo_id}.json"), "w", encoding="utf-8") as f:
        json.dump(grupo, f, ensure_ascii=False)
    STAGING.pop(staging_id, None)
    return grupo


@app.get("/api/diario")
def api_diario_historial():
    """La memoria diaria: todas las conciliaciones diarias guardadas, de la
    más reciente a la más vieja, con su resumen. Los tableros nunca se borran
    y los pendientes encadenan cada día con el siguiente."""
    grupos = []
    for nombre in os.listdir(DATOS_DIR):
        if not (nombre.startswith("grupo_") and nombre.endswith(".json")):
            continue
        ruta = os.path.join(DATOS_DIR, nombre)
        try:
            with open(ruta, encoding="utf-8") as f:
                g = json.load(f)
        except (json.JSONDecodeError, OSError):
            continue
        cuentas = g.get("cuentas", [])
        ok = [c for c in cuentas if c.get("estado") == "ok"]
        r = lambda c, k: (c.get("resumen") or {}).get(k) or 0
        expl = [c["resumen"]["porcentaje_explicado"] for c in ok
                if (c.get("resumen") or {}).get("porcentaje_explicado") is not None]
        fechas_mov = [c.get(k) for c in cuentas for k in ("desde", "hasta") if c.get(k)]
        grupos.append({
            "desde": min(fechas_mov) if fechas_mov else None,
            "hasta": max(fechas_mov) if fechas_mov else None,
            "grupo_id": g.get("grupo_id"),
            "procesado": g.get("procesado"), "hora": g.get("hora"),
            "cuentas": len(cuentas), "conciliadas": len(ok),
            "incompletas": len(cuentas) - len(ok),
            "movimientos": sum(r(c, "movimientos_banco") for c in ok),
            "arrastrados": sum(r(c, "arrastrados") for c in ok),
            "pendientes": sum(r(c, "banco_sin_contabilizar") + r(c, "mayor_sin_banco")
                              for c in ok),
            "explicado_promedio": round(sum(expl) / len(expl), 1) if expl else None,
            "_orden": os.path.getmtime(ruta),
        })
    grupos.sort(key=lambda x: x["_orden"], reverse=True)
    for g in grupos:
        g.pop("_orden")
    return {"grupos": grupos}


@app.get("/api/diario/{grupo_id}")
def api_diario_grupo(grupo_id: str):
    ruta = os.path.join(DATOS_DIR, f"grupo_{grupo_id}.json")
    if not os.path.exists(ruta):
        return JSONResponse(status_code=404, content={"error": "Tablero no encontrado"})
    with open(ruta, encoding="utf-8") as f:
        grupo = json.load(f)
    # refrescar los números con el estado actual de cada job (conciliaciones
    # manuales posteriores incluidas)
    for fila in grupo["cuentas"]:
        if fila.get("job_id"):
            datos = _obtener(fila["job_id"])
            if datos:
                fila["resumen"] = _resumen_mini(datos["resumen"])
    return grupo


@app.delete("/api/diario/{grupo_id}")
def api_diario_eliminar(grupo_id: str):
    """Elimina una corrida diaria completa (p. ej. se subió un extracto
    equivocado): borra el tablero y los jobs de sus cuentas, y rebobina la
    cadena de arrastre de cada cuenta a su corrida anterior."""
    ruta = os.path.join(DATOS_DIR, f"grupo_{grupo_id}.json")
    if not os.path.exists(ruta):
        return JSONResponse(status_code=404, content={"error": "Tablero no encontrado"})
    with open(ruta, encoding="utf-8") as f:
        grupo = json.load(f)

    arrastre = _cargar_arrastre()
    rebobinadas = []
    for fila in grupo.get("cuentas", []):
        job_id = fila.get("job_id")
        if not job_id:
            continue
        datos = _obtener(job_id) or {}
        cid = fila.get("cuenta_id")
        # si esta corrida era el último eslabón del arrastre, volver al anterior
        if cid and arrastre.get(cid) == job_id:
            anterior = datos.get("arrastre_desde") or \
                (datos.get("arrastre") or {}).get("desde_job")
            if anterior and _obtener(anterior):
                arrastre[cid] = anterior
                rebobinadas.append({"cuenta_id": cid, "vuelve_a": anterior})
            else:
                arrastre.pop(cid, None)
                rebobinadas.append({"cuenta_id": cid, "vuelve_a": None})
        RESULTADOS.pop(job_id, None)
        try:
            os.remove(os.path.join(DATOS_DIR, f"{job_id}.json"))
        except OSError:
            pass
    _guardar_arrastre(arrastre)
    os.remove(ruta)
    return {"ok": True, "grupo_id": grupo_id, "arrastre_rebobinado": rebobinadas}


@app.get("/api/gastos")
def api_gastos():
    """Conceptos definidos por el usuario que se clasifican como gasto bancario."""
    return {"gastos": gastos_conf.cargar()}


@app.post("/api/gastos")
def api_gastos_agregar(cuerpo: dict = Body(...)):
    """Agrega un concepto de gasto. Si viene job_id, además reclasifica en ese
    resultado los movimientos de 'banco sin contabilizar' que lo contengan."""
    terminos, error = gastos_conf.agregar(cuerpo.get("termino"))
    if error:
        return JSONResponse(status_code=422, content={"error": error})

    datos, movidos = None, 0
    job_id = cuerpo.get("job_id")
    if job_id:
        datos = _obtener(job_id)
    if datos:
        termino = gastos_conf.limpiar_termino(cuerpo.get("termino"))
        se_mueven = [x for x in datos["banco_sin_contabilizar"]
                     if gastos_conf.es_gasto(x["descripcion"], [{"termino": termino}])]
        if se_mueven:
            ids = {x["id"] for x in se_mueven}
            datos["banco_sin_contabilizar"] = [
                x for x in datos["banco_sin_contabilizar"] if x["id"] not in ids]
            datos["gastos_bancarios"].extend(se_mueven)
            datos["gastos_bancarios"].sort(key=lambda x: x["fecha"] or "")
            movidos = len(se_mueven)
            _recalcular_resumen(datos)
            _guardar(job_id)
    return {"gastos": terminos, "movidos": movidos, "datos": datos}


@app.delete("/api/gastos/{term_id}")
def api_gastos_eliminar(term_id: str):
    return {"gastos": gastos_conf.eliminar(term_id)}


@app.post("/api/excluir/{job_id}")
def api_excluir(job_id: str, cuerpo: dict = Body(...)):
    """Saca asientos del mayor de la conciliación (anulados, contraasientos que
    la detección automática no encontró, etc.). Quedan referenciados y se
    pueden restaurar."""
    datos = _obtener(job_id)
    if not datos:
        return JSONResponse(status_code=404, content={"error": "Resultado no encontrado"})
    ids = set(cuerpo.get("ids_mayor") or [])
    if not ids:
        return JSONResponse(status_code=422, content={"error": "No hay asientos seleccionados"})

    items = []
    for lista in LISTAS_MAYOR:
        for x in datos[lista]:
            if x["id"] in ids:
                items.append({**x, "_origen": lista})
    if len(items) != len(ids):
        return JSONResponse(status_code=409, content={
            "error": "Algunos asientos ya no están disponibles"})
    for lista in LISTAS_MAYOR:
        datos[lista] = [x for x in datos[lista] if x["id"] not in ids]

    datos.setdefault("excluidos_mayor", []).append({
        "excl_id": uuid.uuid4().hex[:10],
        "mayor": items,
        "nota": (cuerpo.get("nota") or "").strip(),
    })
    _recalcular_resumen(datos)
    _guardar(job_id)
    return datos


@app.delete("/api/excluir/{job_id}/{excl_id}")
def api_excluir_deshacer(job_id: str, excl_id: str):
    datos = _obtener(job_id)
    if not datos:
        return JSONResponse(status_code=404, content={"error": "Resultado no encontrado"})
    entrada = next((e for e in datos.get("excluidos_mayor", [])
                    if e["excl_id"] == excl_id), None)
    if not entrada:
        return JSONResponse(status_code=404, content={"error": "Exclusión no encontrada"})
    datos["excluidos_mayor"] = [e for e in datos["excluidos_mayor"]
                                if e["excl_id"] != excl_id]
    for x in entrada["mayor"]:
        origen = x.pop("_origen", "e_sin_banco")
        datos[origen].append(x)
        datos[origen].sort(key=lambda i: i["fecha"] or "")
    _recalcular_resumen(datos)
    _guardar(job_id)
    return datos


@app.get("/api/progreso/{job_id}")
def api_progreso(job_id: str):
    p = PROGRESO.get(job_id)
    if p:
        return p
    if _obtener(job_id):
        return {"estado": "listo", "porcentaje": 100, "eta_seg": 0,
                "fase": "Conciliación terminada"}
    return JSONResponse(status_code=404, content={"error": "Job no encontrado"})


@app.get("/api/resultado/{job_id}")
def api_resultado(job_id: str):
    datos = _obtener(job_id)
    if not datos:
        return JSONResponse(status_code=404, content={"error": "Resultado no encontrado"})
    return datos


@app.post("/api/manual/{job_id}")
def api_conciliar_manual(job_id: str, cuerpo: dict = Body(...)):
    """Registra una conciliación manual (N movimientos del banco ↔ M asientos)."""
    datos = _obtener(job_id)
    if not datos:
        return JSONResponse(status_code=404, content={"error": "Resultado no encontrado"})
    datos.setdefault("conciliados_manual", [])

    ids_banco = set(cuerpo.get("ids_banco") or [])
    ids_mayor = set(cuerpo.get("ids_mayor") or [])
    if not ids_banco or not ids_mayor:
        return JSONResponse(status_code=422, content={
            "error": "Hay que seleccionar al menos un movimiento del banco y un asiento del mayor"})

    # localizar y extraer los ítems de las listas residuales
    items_banco, items_mayor = [], []
    for lista in LISTAS_BANCO:
        for x in datos[lista]:
            if x["id"] in ids_banco:
                items_banco.append({**x, "_origen": lista})
    for lista in LISTAS_MAYOR:
        for x in datos[lista]:
            if x["id"] in ids_mayor:
                items_mayor.append({**x, "_origen": lista})

    encontrados = {x["id"] for x in items_banco} | {x["id"] for x in items_mayor}
    faltantes = (ids_banco | ids_mayor) - encontrados
    if faltantes:
        return JSONResponse(status_code=409, content={
            "error": f"Estos ítems ya no están disponibles (¿ya conciliados?): {sorted(faltantes)}"})

    for lista in LISTAS_BANCO:
        datos[lista] = [x for x in datos[lista] if x["id"] not in ids_banco]
    for lista in LISTAS_MAYOR:
        datos[lista] = [x for x in datos[lista] if x["id"] not in ids_mayor]

    neto_banco = round(sum(x["credito"] - x["debito"] for x in items_banco), 2)
    neto_mayor = round(sum(x["debe"] - x["haber"] for x in items_mayor), 2)
    match = {
        "match_id": uuid.uuid4().hex[:10],
        "banco": items_banco,
        "mayor": items_mayor,
        "neto_banco": neto_banco,
        "neto_mayor": neto_mayor,
        "diferencia": round(neto_banco - neto_mayor, 2),
        "nota": (cuerpo.get("nota") or "").strip(),
        "fuente": cuerpo.get("fuente") or "manual",
    }
    datos["conciliados_manual"].append(match)

    # aprender la regla para futuras conciliaciones
    regla = reglas_mod.aprender(match)
    match["regla_aprendida"] = bool(regla)

    idx = cuerpo.get("sugerencia_idx")
    if idx is not None and 0 <= idx < len(datos["ia"]["sugerencias"]):
        datos["ia"]["sugerencias"][idx]["aceptada"] = True

    _recalcular_resumen(datos)
    _guardar(job_id)
    return datos


@app.post("/api/confirmar/{job_id}")
def api_confirmar(job_id: str, cuerpo: dict = Body(...)):
    """Marca movimientos 'pendientes de confirmar (O)' como confirmados en el
    sistema contable (o los desmarca). Es un checklist de seguimiento."""
    datos = _obtener(job_id)
    if not datos:
        return JSONResponse(status_code=404, content={"error": "Resultado no encontrado"})

    if "todos" in cuerpo:
        val = bool(cuerpo["todos"])
        for m in datos["conciliados_o"]:
            m["confirmado"] = val
    else:
        banco_id = cuerpo.get("banco_id")
        m = next((m for m in datos["conciliados_o"]
                  if m["banco"]["id"] == banco_id), None)
        if not m:
            return JSONResponse(status_code=404, content={"error": "Movimiento no encontrado"})
        m["confirmado"] = bool(cuerpo.get("confirmado", True))

    datos["resumen"]["o_confirmados"] = sum(
        1 for m in datos["conciliados_o"] if m.get("confirmado"))
    _guardar(job_id)
    return datos


@app.delete("/api/manual/{job_id}/{match_id}")
def api_deshacer_manual(job_id: str, match_id: str):
    """Deshace una conciliación manual: los ítems vuelven a sus listas."""
    datos = _obtener(job_id)
    if not datos:
        return JSONResponse(status_code=404, content={"error": "Resultado no encontrado"})
    match = next((m for m in datos.get("conciliados_manual", [])
                  if m["match_id"] == match_id), None)
    if not match:
        return JSONResponse(status_code=404, content={"error": "Match no encontrado"})

    datos["conciliados_manual"] = [m for m in datos["conciliados_manual"]
                                   if m["match_id"] != match_id]
    reglas_mod.olvidar(match)
    for x in match["banco"]:
        origen = x.pop("_origen", "banco_sin_contabilizar")
        datos[origen].append(x)
        datos[origen].sort(key=lambda i: i["fecha"] or "")
    for x in match["mayor"]:
        origen = x.pop("_origen", "e_sin_banco")
        datos[origen].append(x)
        datos[origen].sort(key=lambda i: i["fecha"] or "")

    _recalcular_resumen(datos)
    _guardar(job_id)
    return datos


@app.get("/api/exportar/{job_id}")
def api_exportar(job_id: str):
    datos = _obtener(job_id)
    if not datos:
        return JSONResponse(status_code=404, content={"error": "Resultado no encontrado"})
    buffer = _generar_excel(datos)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="conciliacion.xlsx"'},
    )


def _generar_excel(datos):
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    encabezado = Font(bold=True, color="FFFFFF")
    relleno = PatternFill("solid", fgColor="C8102E")

    def hoja(nombre, columnas, filas):
        ws = wb.create_sheet(nombre)
        ws.append(columnas)
        for c in ws[1]:
            c.font = encabezado
            c.fill = relleno
        for f in filas:
            ws.append(f)
        for col in ws.columns:
            ancho = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(ancho + 2, 60)
        return ws

    # Resumen
    ws = wb.active
    ws.title = "Resumen"
    r = datos["resumen"]
    filas = [
        ("Movimientos del banco", r["movimientos_banco"]),
        ("Conciliados con cuenta E", r["conciliados_e"]),
        ("En cuenta O pendientes de confirmar", r["en_o_pendientes_confirmar"]),
        ("Conciliados manualmente (grupos)", r.get("conciliados_manual", 0)),
        ("Gastos/impuestos bancarios sin contabilizar",
         f'{r["gastos_bancarios"]["cantidad"]}  ($ {r["gastos_bancarios"]["importe"]:,.2f})'),
        ("Otros movimientos del banco sin contabilizar",
         f'{r["banco_sin_contabilizar"]["cantidad"]}  ($ {r["banco_sin_contabilizar"]["importe"]:,.2f})'),
        ("Asientos E sin movimiento en el banco",
         f'{r["e_sin_banco"]["cantidad"]}  ($ {r["e_sin_banco"]["importe"]:,.2f})'),
        ("Pendientes en O sin movimiento en el banco",
         f'{r["o_pendientes_sin_banco"]["cantidad"]}  ($ {r["o_pendientes_sin_banco"]["importe"]:,.2f})'),
        ("% del extracto conciliado asiento por asiento", f'{r["porcentaje_conciliado"]}%'),
        ("% del extracto explicado (incluye gastos justificados por ND mensual)",
         f'{r.get("porcentaje_explicado", r["porcentaje_conciliado"])}%'),
    ]
    ws.append(["Concepto", "Valor"])
    for c in ws[1]:
        c.font = encabezado
        c.fill = relleno
    for f in filas:
        ws.append(f)
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 30

    def fila_match(m):
        b, a = m["banco"], m["asiento"]
        return (b["fecha"], b["comprobante"], b["descripcion"], b["debito"] or "",
                b["credito"] or "", a["hoja"], a["asiento"], a["fecha"],
                a["referencia"], a["comentario"], a["debe"] or "", a["haber"] or "",
                m["metodo"])

    cols_match = ["Fecha banco", "Comprobante", "Descripción banco", "Débito", "Crédito",
                  "Hoja", "Asiento", "Fecha mayor", "Referencia", "Comentario",
                  "Debe", "Haber", "Método"]
    hoja("Conciliados (E)", cols_match, [fila_match(m) for m in datos["conciliados_e"]])
    hoja("Pendientes confirmar (O)", cols_match + ["Confirmado en FES"],
         [fila_match(m) + ("SÍ" if m.get("confirmado") else "",)
          for m in datos["conciliados_o"]])

    cols_banco = ["Fecha", "Comprobante", "Descripción", "Detalle", "Débito", "Crédito", "Archivo"]
    fila_banco = lambda b: (b["fecha"], b["comprobante"], b["descripcion"], b["detalle"],
                            b["debito"] or "", b["credito"] or "", b["archivo"])
    hoja("Banco sin contabilizar", cols_banco,
         [fila_banco(b) for b in datos["banco_sin_contabilizar"]])
    hoja("Gastos bancarios", cols_banco,
         [fila_banco(b) for b in datos["gastos_bancarios"]])

    cols_mayor = ["Hoja", "Asiento", "Fecha", "Referencia", "Comentario", "Debe", "Haber"]
    fila_mayor = lambda a: (a["hoja"], a["asiento"], a["fecha"], a["referencia"],
                            a["comentario"], a["debe"] or "", a["haber"] or "")
    hoja("Mayor E sin banco", cols_mayor, [fila_mayor(a) for a in datos["e_sin_banco"]])
    hoja("O pendientes sin banco", cols_mayor,
         [fila_mayor(a) for a in datos["o_pendientes_sin_banco"]])

    filas_x = []
    for p in datos.get("contraasientos", []):
        filas_x.append(fila_mayor(p["debe"]) + ("contraasiento (automático)",))
        filas_x.append(fila_mayor(p["haber"]) + ("contraasiento (automático)",))
    for e in datos.get("excluidos_mayor", []):
        motivo = "excluido manualmente" + (f' — {e["nota"]}' if e["nota"] else "")
        for a in e["mayor"]:
            filas_x.append(fila_mayor(a) + (motivo,))
    hoja("Contraasientos y excluidos", cols_mayor + ["Motivo"], filas_x)

    # Nota de débito mensual (gastos bancarios agrupados)
    filas_nd = []
    for p in datos["nota_debito"]:
        for c in p["categorias"]:
            obs = ""
            if "iva_esperado" in c:
                dif = round(c["importe"] - c["iva_esperado"], 2)
                obs = f'IVA esperado $ {c["iva_esperado"]:,.2f} (dif {dif:,.2f})'
            filas_nd.append((p["periodo"], c["categoria"], c["cantidad"], c["importe"], obs))
        filas_nd.append((p["periodo"], "TOTAL SEGÚN EXTRACTO", "", p["total_extracto"], ""))
        for a in p["asientos_mayor"]:
            lado = a["haber"] or -a["debe"]
            filas_nd.append((p["periodo"], f'Mayor [{a["hoja"]}] {a["referencia"]} — {a["comentario"]}',
                             "", lado, "asiento del mayor"))
        if p["total_mayor"] is not None:
            filas_nd.append((p["periodo"], "TOTAL SEGÚN MAYOR", "", p["total_mayor"], ""))
            filas_nd.append((p["periodo"], "DIFERENCIA (extracto - mayor)", "",
                             p["diferencia"], ""))
        filas_nd.append(("", "", "", "", ""))
    hoja("Gastos por mes (ND)",
         ["Período", "Concepto", "Cant.", "Importe", "Observación"], filas_nd)

    # Detalle de conceptos por período (para investigar diferencias)
    filas_con = []
    for p in datos["nota_debito"]:
        for c in p["conceptos"]:
            filas_con.append((p["periodo"], c["descripcion"], c["cantidad"], c["importe"]))
    hoja("Gastos detalle conceptos",
         ["Período", "Concepto del extracto", "Cant.", "Importe"], filas_con)

    # Conciliados manualmente (grupos N a M)
    filas_man = []
    for n, m in enumerate(datos.get("conciliados_manual", []), 1):
        for b in m["banco"]:
            filas_man.append((n, "BANCO", b["fecha"], b["comprobante"],
                              f'{b["descripcion"]} {b["detalle"]}'.strip(),
                              b["debito"] or "", b["credito"] or "", "", m["nota"]))
        for a in m["mayor"]:
            filas_man.append((n, f'MAYOR {a["hoja"]}', a["fecha"], a["asiento"],
                              f'{a["referencia"]} — {a["comentario"]}',
                              a["haber"] or "", a["debe"] or "", "", ""))
        filas_man.append((n, "→ diferencia", "", "", "", "", "", m["diferencia"],
                          "aceptada de IA" if m["fuente"] == "ia" else ""))
    hoja("Conciliados manualmente",
         ["Grupo", "Lado", "Fecha", "Comp/Asiento", "Descripción",
          "Débito/Haber", "Crédito/Debe", "Diferencia", "Nota"], filas_man)

    if datos["ia"]["sugerencias"]:
        hoja("Sugerencias IA",
             ["IDs banco", "IDs mayor", "Confianza", "Motivo"],
             [(", ".join(s["ids_banco"]), ", ".join(s["ids_mayor"]),
               s["confianza"], s["motivo"]) for s in datos["ia"]["sugerencias"]])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@app.get("/diario")
def pagina_diario():
    return FileResponse(os.path.join(BASE_DIR, "static", "diario.html"),
                        headers={"Cache-Control": "no-cache"})


@app.get("/")
def index():
    # sin caché: tras cada deploy el navegador debe traer la interfaz nueva
    return FileResponse(os.path.join(BASE_DIR, "static", "index.html"),
                        headers={"Cache-Control": "no-cache"})


app.mount("/static", StaticFiles(directory=os.path.join(BASE_DIR, "static")), name="static")


if __name__ == "__main__":
    import uvicorn
    puerto = int(os.environ.get("PORT", "8765"))
    host = "0.0.0.0" if os.environ.get("PORT") else "127.0.0.1"
    uvicorn.run(app, host=host, port=puerto)
